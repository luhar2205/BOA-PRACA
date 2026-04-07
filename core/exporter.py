import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def salvar_excel_kildere(df, caminho):
    df.to_excel(caminho, index=False)
    wb = load_workbook(caminho)
    ws = wb.active

    FILL_HDR = PatternFill("solid", fgColor="1F4068") 
    FILL_UA = PatternFill("solid", fgColor="D9EAD3") 
    FONT_W = Font(bold=True, color="FFFFFF")
    BORDA = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill, cell.font, cell.border = FILL_HDR, FONT_W, BORDA
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value
            cell.border = BORDA
            if "VLR UNIT" in str(col_name): cell.number_format = 'R$ #,##0.00'
            if col_name in ["CÓD SANKHYA", "UA / CONVERSÃO"]: cell.fill = FILL_UA
            cell.alignment = Alignment(vertical="center", horizontal="left")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 3, 55)

    wb.save(caminho)
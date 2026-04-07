import os
import shutil
import uuid
import zipfile
import re
from datetime import datetime
import xml.etree.ElementTree as ET

import pandas as pd
from thefuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from fastapi import FastAPI, File, UploadFile, BackgroundTasks
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
import uvicorn

app = FastAPI(title="Boa Praça OS - Enterprise")

# =====================================================================
# 🛡️ 1. CONFIGURAÇÕES DE DIRETÓRIO
# =====================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMP_DIR = os.path.join(BASE_DIR, "temp_files")
os.makedirs(TEMP_DIR, exist_ok=True)
CAMINHO_BASE_SANKHYA = os.path.join(BASE_DIR, "PRODUTOS_KILDERE.xlsx")

def limpar_pasta_temporaria(caminho_pasta: str):
    try:
        if os.path.exists(caminho_pasta):
            shutil.rmtree(caminho_pasta)
    except Exception: pass

# =====================================================================
# ⚙️ 2. MOTOR DE INTELIGÊNCIA E MATEMÁTICA (Tudo Embutido)
# =====================================================================
def calcular_peso_pacote_unitario(desc):
    if not desc: return None
    desc = desc.upper().replace(',', '.')
    match_mult = re.search(r'(\d+)\s*X\s*(\d+[.]?\d*)\s*(KG|G|GR)', desc)
    if match_mult:
        return (float(match_mult.group(1)) * float(match_mult.group(2))) if match_mult.group(3) == 'KG' else (float(match_mult.group(1)) * float(match_mult.group(2))) / 1000
    match_simples = re.search(r'(\d+[.]?\d*)\s*(KG|G|GR)', desc)
    if match_simples:
        return float(match_simples.group(1)) if match_simples.group(2) == 'KG' else float(match_simples.group(1)) / 1000
    return None

def calcular_ua_final_blindada(desc_xml, desc_sankhya, q_com, q_trib):
    if q_com <= 0: return "0.000000"
    peso_pacote = calcular_peso_pacote_unitario(desc_sankhya) or calcular_peso_pacote_unitario(desc_xml)
    ua_bruta = q_trib / q_com
    if peso_pacote and peso_pacote > 0:
        ua_real = ua_bruta / peso_pacote
        peso_str = f"{peso_pacote*1000:.0f}g" if peso_pacote < 1 else f"{peso_pacote:.1f}kg"
        return f"{ua_real:.6f} (Ref: {round(ua_real)} un de {peso_str}/cx)"
    return f"{ua_bruta:.6f}"

def extrair_dados_xml(caminho):
    try:
        tree = ET.parse(caminho)
        root = tree.getroot()
        for elem in root.iter():
            if "}" in elem.tag: elem.tag = elem.tag.split("}", 1)[1]
        
        num_nf = root.find(".//ide/nNF").text
        emitente = root.find(".//emit/xNome").text
        
        itens = []
        for det in root.findall(".//det"):
            prod = det.find("prod")
            ean_trib = prod.find("cEANTrib").text if prod.find("cEANTrib") is not None else ""
            ean_com = prod.find("cEAN").text if prod.find("cEAN") is not None else ""
            ean_final = ean_trib if ean_trib not in ["SEM GTIN", ""] else ean_com
            itens.append({
                "NF": num_nf, "FORNECEDOR": emitente,
                "Produto_XML": prod.find("xProd").text,
                "UN_NOTA": prod.find("uCom").text if prod.find("uCom") is not None else "",
                "EAN_XML": ean_final,
                "QTD_NOTA": float(prod.find("qCom").text or 0),
                "VLR_UNIT_NOTA": float(prod.find("vUnCom").text or 0),
                "qTrib": float(prod.find("qTrib").text or 0)
            })
        return itens
    except: return []

def _expand_abbreviations(text):
    replaces = {r'\bF\.\b': 'FILE', r'\bF\b': 'FILE', r'\bSALMON\b': 'SALMAO', r'\bDEF\b': 'DEFUMADO', r'\bS/E\b': 'SEM ESPINHA', r'\bS/O\b': 'SEM OSSO', r'\bC/O\b': 'COM OSSO', r'\bCXA\b': 'CAIXA', r'\bCX\b': 'CAIXA', r'\bPCT\b': 'PACOTE', r'\bPT\b': 'PACOTE', r'\bRESF\b': 'RESFRIADO', r'\BCONG\b': 'CONGELADO', r'\bLING\b': 'LINGUICA', r'\bCALAB\b': 'CALABRESA', r'\bPREM\b': 'PREMIUM', r'\bESP\b': 'ESPECIAL'}
    for pattern, replacement in replaces.items(): text = re.sub(pattern, replacement, text)
    return text

def _norm(v): 
    if pd.isna(v): return ""
    return _expand_abbreviations(str(v).strip().upper())

def _clean_ean(ean):
    e = str(ean).strip()
    if e in ["", "SEM GTIN", "N/A", "NONE"]: return ""
    e = e.lstrip('0') 
    if len(e) == 14 and e.startswith('1'): e = e[1:] 
    return e

def rodar_automacao_total(pasta_xml, arquivo_base):
    df_raw = pd.read_excel(arquivo_base, header=2).fillna("")
    df_raw.columns = [c.strip() for c in df_raw.columns]
    ean_map = { _clean_ean(row.get('Referência', '')): i for i, row in df_raw.iterrows() if _clean_ean(row.get('Referência', '')) }
    
    xmls = list(os.path.join(pasta_xml, f) for f in os.listdir(pasta_xml) if f.lower().endswith('.xml'))
    todos_itens = []
    for p in xmls: todos_itens.extend(extrair_dados_xml(p))

    final = []
    for it in todos_itens:
        match_idx = None
        tipo_match = "❌ SEM MATCH"
        desc_xml_raw = str(it["Produto_XML"]).strip().upper()
        desc_xml_norm = _norm(it["Produto_XML"]) 
        ean_xml = _clean_ean(it["EAN_XML"]) 

        if ean_xml and ean_xml in ean_map:
            match_idx = ean_map[ean_xml]
            tipo_match = "✅ EAN"
        
        if match_idx is None:
            scores = []
            for i, row in df_raw.iterrows():
                desc_s_norm = _norm(row.get('Descrição', ''))
                marca_s = _norm(row.get('Marca', ''))
                score_base = fuzz.token_set_ratio(desc_xml_norm, desc_s_norm)
                if marca_s and marca_s in desc_xml_norm: score_base += 20
                if any(x in desc_xml_norm for x in ["500G", "1KG", "0.5KG"]) and any(x in desc_s_norm for x in ["500G", "1KG", "0.5KG"]): score_base += 15
                if score_base > 75: scores.append((i, score_base))
            if scores:
                scores.sort(key=lambda x: x[1], reverse=True)
                match_idx = scores[0][0]
                tipo_match = f"🔶 INTELLIGENT ({scores[0][1]}%)"

        row_s = df_raw.iloc[match_idx] if match_idx is not None else {}
        desc_s_final = str(row_s.get('Descrição', "")).strip().upper()
        ua = calcular_ua_final_blindada(desc_xml_raw, desc_s_final, it["QTD_NOTA"], it["qTrib"])

        final.append({
            "NF": it["NF"], "FORNECEDOR": it["FORNECEDOR"], "EAN XML": it["EAN_XML"], 
            "PRODUTO XML": desc_xml_raw, "UN NOTA": it["UN_NOTA"], "QTD NOTA": it["QTD_NOTA"],
            "VLR UNIT NOTA": it["VLR_UNIT_NOTA"], "CÓD SANKHYA": row_s.get('Código', ""),
            "DESC SANKHYA": desc_s_final, "UA / CONVERSÃO": ua, "MATCH": tipo_match
        })
    return pd.DataFrame(final)

def salvar_excel_kildere(df, caminho):
    df.to_excel(caminho, index=False)
    wb = load_workbook(caminho)
    ws = wb.active
    FILL_HDR = PatternFill("solid", fgColor="1F4068") 
    FILL_UA = PatternFill("solid", fgColor="D9EAD3") 
    FONT_W = Font(bold=True, color="FFFFFF")
    BORDA = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.fill, cell.font, cell.border, cell.alignment = FILL_HDR, FONT_W, BORDA, Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value
            cell.border, cell.alignment = BORDA, Alignment(vertical="center", horizontal="left")
            if "VLR UNIT" in str(col_name): cell.number_format = 'R$ #,##0.00'
            if col_name in ["CÓD SANKHYA", "UA / CONVERSÃO"]: cell.fill = FILL_UA
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 55)
    wb.save(caminho)

# =====================================================================
# 🎨 3. FRONT-END (HTML / TELA)
# =====================================================================
HTML_CONTENT = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Boa Praça OS | Hub Operacional</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <style>
        body { font-family: sans-serif; background-color: #f3f4f6; }
        .drag-active { border-color: #3b82f6 !important; background-color: #eff6ff !important; }
    </style>
</head>
<body class="h-screen flex overflow-hidden text-gray-800">
    <aside class="w-64 bg-[#1F4068] text-white flex flex-col shadow-2xl z-20">
        <div class="p-6 border-b border-blue-800/50">
            <h1 class="text-2xl font-extrabold flex items-center"><i class="fa-solid fa-anchor mr-3 text-blue-400"></i> BOA PRAÇA</h1>
        </div>
        <nav class="flex-1 p-4 space-y-2">
            <a href="#" class="flex items-center p-3 bg-blue-800/50 text-white rounded-lg font-semibold"><i class="fa-solid fa-microchip w-6 text-blue-400"></i> Motor Sankhya UA</a>
        </nav>
    </aside>

    <main class="flex-1 flex flex-col h-screen overflow-hidden bg-gray-50 relative">
        <header class="h-16 bg-white shadow-sm flex items-center justify-between px-8 z-10">
            <h2 class="text-xl font-bold text-gray-700">Processamento Inteligente de Notas Fiscais</h2>
        </header>

        <div class="flex-1 overflow-y-auto p-8">
            <div id="uploadScreen" class="max-w-4xl mx-auto mt-4">
                <div class="bg-white rounded-2xl shadow-xl p-10 border border-gray-100 text-center">
                    <h3 class="text-2xl font-extrabold text-[#1F4068]">Importar Lote (XML/ZIP)</h3>
                    <form id="uploadForm" class="mt-8">
                        <div id="dropzone" class="border-2 border-dashed border-gray-300 rounded-xl p-16 relative cursor-pointer hover:bg-blue-50">
                            <input type="file" id="fileInput" accept=".xml, .zip" multiple required class="absolute inset-0 w-full h-full opacity-0 cursor-pointer z-50"/>
                            <i class="fa-solid fa-cloud-arrow-up text-5xl text-blue-500 mb-4"></i>
                            <h3 class="font-bold text-lg text-gray-700">Arraste os arquivos para cá</h3>
                            <p id="file-names" class="text-sm text-blue-600 font-bold mt-4"></p>
                        </div>
                        <button type="submit" id="submitBtn" class="w-full bg-[#1F4068] hover:bg-[#162d4a] text-white font-bold py-4 rounded-xl shadow-lg mt-6">
                            PROCESSAR NOTAS
                        </button>
                    </form>
                </div>
            </div>

            <div id="loadingScreen" class="hidden h-full flex flex-col items-center justify-center pt-20">
                <div class="w-16 h-16 border-4 border-gray-200 border-t-[#1F4068] rounded-full animate-spin"></div>
                <h2 class="text-2xl font-bold text-[#1F4068] mt-6">Processando Inteligência...</h2>
            </div>

            <div id="resultScreen" class="hidden max-w-7xl mx-auto">
                <div class="flex justify-between items-end mb-6">
                    <h2 class="text-3xl font-extrabold text-[#1F4068]">Relatório de Operação</h2>
                    <div class="flex space-x-3">
                        <button onclick="window.location.reload()" class="px-6 py-2 bg-white border border-gray-300 font-bold rounded-lg hover:bg-gray-50">Novo Lote</button>
                        <a id="downloadBtn" href="#" class="px-6 py-2 bg-green-600 hover:bg-green-700 text-white font-bold rounded-lg">Baixar Excel</a>
                    </div>
                </div>

                <div class="bg-white rounded-xl shadow-sm border border-gray-200 overflow-hidden">
                    <div class="overflow-x-auto max-h-[500px]">
                        <table class="w-full text-left text-sm" id="dataTable">
                            <thead class="bg-gray-100 text-gray-600 sticky top-0 shadow-sm">
                                <tr>
                                    <th class="p-4 font-bold">Cód</th>
                                    <th class="p-4 font-bold">Descrição Base</th>
                                    <th class="p-4 font-bold">Produto XML</th>
                                    <th class="p-4 font-bold text-center bg-blue-50 text-blue-800">Cálculo UA</th>
                                    <th class="p-4 font-bold text-center">Status</th>
                                </tr>
                            </thead>
                            <tbody id="tableBody" class="divide-y divide-gray-100"></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>
    </main>

    <script>
        const fileInput = document.getElementById('fileInput');
        const dropzone = document.getElementById('dropzone');
        const fileNames = document.getElementById('file-names');
        
        fileInput.addEventListener('dragenter', () => dropzone.classList.add('drag-active'));
        fileInput.addEventListener('dragleave', () => dropzone.classList.remove('drag-active'));
        fileInput.addEventListener('drop', () => dropzone.classList.remove('drag-active'));
        fileInput.addEventListener('change', () => { fileNames.textContent = Array.from(fileInput.files).map(f => f.name).join(', '); });

        document.getElementById('uploadForm').addEventListener('submit', async (e) => {
            e.preventDefault();
            document.getElementById('uploadScreen').classList.add('hidden');
            document.getElementById('loadingScreen').classList.remove('hidden');

            const formData = new FormData();
            for (let i = 0; i < fileInput.files.length; i++) formData.append('arquivos', fileInput.files[i]);

            try {
                const response = await fetch('/api/processar', { method: 'POST', body: formData });
                const result = await response.json();
                
                if(!response.ok) { alert(result.error); window.location.reload(); return; }

                document.getElementById('downloadBtn').href = `/api/download/${result.session_id}`;
                const tbody = document.getElementById('tableBody');
                
                result.dados.forEach(item => {
                    const row = document.createElement('tr');
                    row.innerHTML = `
                        <td class="p-4 font-bold">${item['CÓD SANKHYA'] || '-'}</td>
                        <td class="p-4 font-bold text-[#1F4068]">${item['DESC SANKHYA']}</td>
                        <td class="p-4 text-xs text-gray-500">${item['PRODUTO XML']}</td>
                        <td class="p-4 text-center bg-blue-50/50 font-mono font-bold">${item['UA / CONVERSÃO']}</td>
                        <td class="p-4 text-center text-xs font-bold">${item['MATCH']}</td>
                    `;
                    tbody.appendChild(row);
                });

                document.getElementById('loadingScreen').classList.add('hidden');
                document.getElementById('resultScreen').classList.remove('hidden');
            } catch (error) { alert("Erro de comunicação."); window.location.reload(); }
        });
    </script>
</body>
</html>
"""

# =====================================================================
# 🚀 4. ROTAS DA API
# =====================================================================
@app.get("/", response_class=HTMLResponse)
async def home(): return HTMLResponse(content=HTML_CONTENT)

@app.post("/api/processar")
async def api_processar(background_tasks: BackgroundTasks, arquivos: list[UploadFile] = File(...)):
    if not os.path.exists(CAMINHO_BASE_SANKHYA):
        return JSONResponse(status_code=500, content={"error": "Base PRODUTOS_KILDERE.xlsx não encontrada no servidor."})

    sessao_id = uuid.uuid4().hex
    pasta_sessao = os.path.join(TEMP_DIR, sessao_id)
    pasta_xmls = os.path.join(pasta_sessao, "xmls")
    os.makedirs(pasta_xmls, exist_ok=True)
        
    for arquivo in arquivos:
        extensao = os.path.splitext(arquivo.filename)[1].lower()
        caminho_temp = os.path.join(pasta_sessao, arquivo.filename)
        with open(caminho_temp, "wb") as buffer: shutil.copyfileobj(arquivo.file, buffer)
            
        if extensao == '.zip':
            try:
                with zipfile.ZipFile(caminho_temp, 'r') as zip_ref:
                    for file_info in zip_ref.infolist():
                        if file_info.filename.lower().endswith('.xml'): zip_ref.extract(file_info, pasta_xmls)
            except: pass
        elif extensao == '.xml':
            shutil.move(caminho_temp, os.path.join(pasta_xmls, arquivo.filename))
            
    df_resultado = rodar_automacao_total(pasta_xmls, CAMINHO_BASE_SANKHYA)
    nome_arquivo = f"Conversao_Sankhya_{datetime.now().strftime('%d_%m_%Hh%M')}.xlsx"
    caminho_excel = os.path.join(pasta_sessao, nome_arquivo)
    salvar_excel_kildere(df_resultado, caminho_excel)
    
    return JSONResponse(content={"session_id": f"{sessao_id}/{nome_arquivo}", "dados": df_resultado.to_dict(orient="records")})

@app.get("/api/download/{sessao_id}/{nome_arquivo}")
async def api_download(sessao_id: str, nome_arquivo: str, background_tasks: BackgroundTasks):
    pasta_sessao = os.path.join(TEMP_DIR, sessao_id)
    background_tasks.add_task(limpar_pasta_temporaria, pasta_sessao)
    return FileResponse(path=os.path.join(pasta_sessao, nome_arquivo), filename=nome_arquivo, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))